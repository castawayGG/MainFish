import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from models.account import Account
from models.proxy import Proxy
from models.admin_log import AdminLog
from datetime import datetime

class ExcelExporter:
    """
    Класс для экспорта данных в Excel-файлы.
    Поддерживает экспорт аккаунтов, прокси, логов и других сущностей.
    """
    
    @staticmethod
    def export_accounts(accounts: list[Account]) -> io.BytesIO:
        """
        Создаёт Excel-файл со списком аккаунтов.
        Возвращает BytesIO с содержимым файла.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Accounts"
        
        # Заголовки
        headers = [
            "ID", "Phone", "Username", "First Name", "Last Name",
            "Premium", "Status", "Created At", "Last Used", "Owner", "Notes"
        ]
        ws.append(headers)
        
        # Стиль заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Данные
        for account in accounts:
            row = [
                account.id,
                account.phone,
                account.username or "",
                account.first_name or "",
                account.last_name or "",
                "Yes" if account.premium else "No",
                account.status,
                account.created_at.strftime("%Y-%m-%d %H:%M") if account.created_at else "",
                account.last_used.strftime("%Y-%m-%d %H:%M") if account.last_used else "",
                account.owner.username if account.owner else "",
                account.notes or ""
            ]
            ws.append(row)
        
        # Автоподбор ширины колонок
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # Сохраняем в BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def export_proxies(proxies: list[Proxy]) -> io.BytesIO:
        """
        Экспорт прокси в Excel.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Proxies"
        
        headers = [
            "ID", "Type", "Host", "Port", "Username", "Password",
            "Status", "Speed (ms)", "Avg Speed", "Enabled", "Last Check", "Last Error",
            "Requests", "Success", "Fails", "Country", "Description", "Created"
        ]
        ws.append(headers)
        
        # Стиль заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        for proxy in proxies:
            row = [
                proxy.id,
                proxy.type,
                proxy.host,
                proxy.port,
                proxy.username or "",
                proxy.password or "",
                proxy.status,
                proxy.speed or "",
                proxy.avg_speed or "",
                "Yes" if proxy.enabled else "No",
                proxy.last_check.strftime("%Y-%m-%d %H:%M") if proxy.last_check else "",
                proxy.last_error or "",
                proxy.requests_count,
                proxy.success_count,
                proxy.fail_count,
                proxy.country or "",
                proxy.description or "",
                proxy.created_at.strftime("%Y-%m-%d %H:%M") if proxy.created_at else ""
            ]
            ws.append(row)
        
        # Автоподбор ширины
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def export_logs(logs: list[AdminLog]) -> io.BytesIO:
        """
        Экспорт логов действий администраторов.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Admin Logs"
        
        headers = ["ID", "Username", "Action", "Details", "IP", "User Agent", "Timestamp"]
        ws.append(headers)
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        for log in logs:
            row = [
                log.id,
                log.username,
                log.action,
                log.details or "",
                log.ip,
                log.user_agent or "",
                log.timestamp.strftime("%Y-%m-%d %H:%M:%S") if log.timestamp else ""
            ]
            ws.append(row)
        
        # Автоподбор ширины
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output